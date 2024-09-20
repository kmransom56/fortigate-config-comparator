
#!/home/keransom/anaconda3/bin/python
##### Begin of Initialize_Teachable_Agent_Skill #####
import openai
from autogen import ConversableAgent, config_list_from_json
from autogen.agentchat.contrib.capabilities.teachability import Teachability


# Begin of Initialize_Teachable_Agent_Skill
def InitializeTeachableAgent():
       print("Entering InitializeTeachableAgent function")
       config_list = config_list_from_json(env_or_file="/home/keransom/OAI_CONFIG_LIST")
       print(f"Config list loaded: {config_list}")
       
       # Set the API key directly in the openai module
       openai.api_key = config_list[0]["api_key"]
       
       llm_config = {
           "config_list": config_list,
           "timeout": 120,
       }
       
       print(f"LLM config created: {llm_config}")
       
       # Initialize the teachable agent
       teachable_agent = ConversableAgent(
           name="FirewallConfigAgent",
           llm_config=llm_config,
       )
       print(f"Teachable agent initialized: {teachable_agent}")
       
       # Enable teachability
       teachability = Teachability(
           reset_db=False,
           path_to_db_dir="./teachability_db"
       )
       print(f"Teachability object created: {teachability}")
       teachability.add_to_agent(teachable_agent)
       print("Teachability added to agent")

       print("Exiting InitializeTeachableAgent function")
       return teachable_agent

#### End of Initialize_Teachable_Agent_Skill ####       

##### Begin of Parse_Network_Requirements_Skill #####

def ParseNetworkRequirements(teachable_agent, document_text):
    # Process the document with the teachable agent
    parsed_data = teachable_agent.process_text(document_text)
    
    # Manually structure the parsed data to match the expected output
    parsed_output = {
        "firewall_policies": parsed_data.get('firewall_policies', []),
        "vlans": parsed_data.get('vlans', []),
        "address_objects": parsed_data.get('address_objects', []),
        "service_objects": parsed_data.get('service_objects', []),
        "web_filters": parsed_data.get('web_filters', []),
        "meraki_switches": parsed_data.get('meraki_switches', []),
        "meraki_stack_routes": parsed_data.get('meraki_stack_routes', [])
    }
    
    return parsed_output

#### End of Parse_Network_Requirements_Skill ####

##### Begin of Transform_To_Excel_Skill #####

from openpyxl import load_workbook

def TransformToExcel(parsed_data, workbook_path):
    # Load the existing workbook
    workbook = load_workbook(workbook_path)
    
    # Populate 'Policy Changes' sheet
    if 'Policy Changes' in workbook.sheetnames:
        sheet = workbook['Policy Changes']
        for i, policy in enumerate(parsed_data.get("firewall_policies", []), start=2):  # Adjust start row as necessary
            sheet.cell(row=i, column=1, value=policy['id'])
            sheet.cell(row=i, column=2, value=policy['srcintf'])
            sheet.cell(row=i, column=3, value=policy['dstintf'])
            sheet.cell(row=i, column=4, value=policy['action'])
    
    # Populate 'VLan' sheet
    if 'VLan' in workbook.sheetnames:
        sheet = workbook['VLan']
        for i, vlan in enumerate(parsed_data.get("vlans", []), start=2):
            sheet.cell(row=i, column=1, value=vlan)
    
    # Populate 'Address Objects' sheet
    if 'Address Objects' in workbook.sheetnames:
        sheet = workbook['Address Objects']
        for i, address in enumerate(parsed_data.get("address_objects", []), start=2):
            sheet.cell(row=i, column=1, value=address['name'])
            sheet.cell(row=i, column=2, value=address['subnet'])
            sheet.cell(row=i, column=3, value=address['description'])
    
    # Populate 'Service Objects' sheet
    if 'Service Objects' in workbook.sheetnames:
        sheet = workbook['Service Objects']
        for i, service in enumerate(parsed_data.get("service_objects", []), start=2):
            sheet.cell(row=i, column=1, value=service['name'])
            sheet.cell(row=i, column=2, value=service['protocol'])
            sheet.cell(row=i, column=3, value=service['port_range'])
    
    # Populate 'WebFilter' sheet
    if 'WebFilter' in workbook.sheetnames:
        sheet = workbook['WebFilter']
        for i, filter in enumerate(parsed_data.get("web_filters", []), start=2):
            sheet.cell(row=i, column=1, value=filter['name'])
            sheet.cell(row=i, column=2, value=filter['category'])
            sheet.cell(row=i, column=3, value=filter['action'])
    
    # Populate 'Meraki Switch' sheet
    if 'Meraki Switch' in workbook.sheetnames:
        sheet = workbook['Meraki Switch']
        for i, switch in enumerate(parsed_data.get("meraki_switches", []), start=2):
            sheet.cell(row=i, column=1, value=switch['name'])
            sheet.cell(row=i, column=2, value=switch['ip'])
            sheet.cell(row=i, column=3, value=switch['port'])
    
    # Populate 'Meraki Stack Routes' sheet
    if 'Meraki Stack Routes' in workbook.sheetnames:
        sheet = workbook['Meraki Stack Routes']
        for i, route in enumerate(parsed_data.get("meraki_stack_routes", []), start=2):
            sheet.cell(row=i, column=1, value=route['destination'])
            sheet.cell(row=i, column=2, value=route['next_hop'])
            sheet.cell(row=i, column=3, value=route['description'])
    
    # Save the modified workbook
    workbook.save("updated_network_requirements.xlsx")

    return workbook

#### End of Transform_To_Excel_Skill ####

##### Begin of Generate_FortiGate_Config_Skill #####

def GenerateFortiGateConfig(parsed_data):
    config_commands = []
    for policy in parsed_data.get("firewall_policies", []):
        command = f"config firewall policy\nedit {policy['id']}\nset srcintf {policy['srcintf']}\nset dstintf {policy['dstintf']}\nset action {policy['action']}\nnext\nend"
        config_commands.append(command)

    with open("fortigate_config.txt", "w") as f:
        f.write("\n".join(config_commands))

    return config_commands

#### End of Generate_FortiGate_Config_Skill ####
        
##### Begin of Reinitialize_Run_Workflow_Skill #####

from autogen import ConversableAgent, config_list_from_json

def ReinitializeAndRunWorkflow(document_text):
    try:
        # Step 1: Reinitialize the LLM client and teachable agent
        config_list = config_list_from_json(env_or_file="OAI_CONFIG_LIST")
        llm_config = {"config_list": config_list, "timeout": 120}

        teachable_agent = ConversableAgent(
            name="FirewallConfigAgent",
            llm_config=llm_config
        )

        # Ensure the client is initialized
        if teachable_agent.llm_config is None:
            raise ValueError("LLM client initialization failed. Please check your configuration.")

        # Step 2: Execute the workflow steps
        parsed_data = ParseNetworkRequirements(teachable_agent, document_text)
        TransformToExcel(parsed_data, "network_requirements.xlsx")
        GenerateFortiGateConfig(parsed_data)

        return "Workflow Completed Successfully"

    except Exception as e:
        return f"Error in running workflow: {e}"

#### End of Reinitialize_Run_Workflow_Skill ####
if __name__ == "__main__":
    print("Main execution block reached")
    teachable_agent = InitializeTeachableAgent()
    print(f"Teachable agent created: {teachable_agent}")
    
    # Test interaction
    response = teachable_agent.generate_reply("Human", "Hello, can you hear me?")
    print(f"Agent response: {response}")
    
    print("Script execution completed")
        